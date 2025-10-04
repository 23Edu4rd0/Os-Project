"""
Módulo para exportação de dados para Excel e CSV.

Suporta exportação de:
- Pedidos/Ordens de Serviço
- Produtos
- Clientes
- Gastos
"""

import csv
import os
from datetime import datetime
from typing import List, Tuple, Optional
from pathlib import Path


class DataExporter:
    """Exportador de dados para Excel e CSV"""
    
    @staticmethod
    def get_export_dir() -> str:
        """
        Retorna o diretório padrão para exportações.
        
        Returns:
            Caminho do diretório de exportações
        """
        export_dir = os.path.expanduser("~/Documents/OrdemServico/Exportacoes")
        os.makedirs(export_dir, exist_ok=True)
        return export_dir
    
    @staticmethod
    def _generate_filename(prefix: str, extension: str) -> str:
        """
        Gera nome de arquivo com timestamp.
        
        Args:
            prefix: Prefixo do arquivo (ex: "pedidos", "clientes")
            extension: Extensão do arquivo (ex: "csv", "xlsx")
            
        Returns:
            Nome do arquivo completo
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
    
    @staticmethod
    def export_to_csv(data: List[Tuple], headers: List[str], filename: str = None, 
                     prefix: str = "export") -> Tuple[bool, str]:
        """
        Exporta dados para CSV.
        
        Args:
            data: Lista de tuplas com os dados
            headers: Lista com nomes das colunas
            filename: Nome do arquivo (opcional)
            prefix: Prefixo do arquivo se filename não for fornecido
            
        Returns:
            Tupla (sucesso: bool, caminho_ou_mensagem_erro: str)
        """
        try:
            # Gerar nome do arquivo se não fornecido
            if not filename:
                filename = DataExporter._generate_filename(prefix, "csv")
            
            # Garantir que termina com .csv
            if not filename.endswith('.csv'):
                filename += '.csv'
            
            # Caminho completo
            export_dir = DataExporter.get_export_dir()
            filepath = os.path.join(export_dir, filename)
            
            # Escrever CSV
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile, delimiter=';', quotechar='"', 
                                  quoting=csv.QUOTE_MINIMAL)
                
                # Escrever cabeçalho
                writer.writerow(headers)
                
                # Escrever dados
                for row in data:
                    # Converter None para string vazia
                    clean_row = [str(val) if val is not None else '' for val in row]
                    writer.writerow(clean_row)
            
            return True, filepath
            
        except Exception as e:
            return False, f"Erro ao exportar CSV: {str(e)}"
    
    @staticmethod
    def export_to_excel(data: List[Tuple], headers: List[str], filename: str = None,
                       prefix: str = "export", sheet_name: str = "Dados") -> Tuple[bool, str]:
        """
        Exporta dados para Excel (.xlsx).
        Requer openpyxl instalado: pip install openpyxl
        
        Args:
            data: Lista de tuplas com os dados
            headers: Lista com nomes das colunas
            filename: Nome do arquivo (opcional)
            prefix: Prefixo do arquivo se filename não for fornecido
            sheet_name: Nome da planilha
            
        Returns:
            Tupla (sucesso: bool, caminho_ou_mensagem_erro: str)
        """
        try:
            # Tentar importar openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                return False, "Biblioteca openpyxl não instalada. Execute: pip install openpyxl"
            
            # Gerar nome do arquivo se não fornecido
            if not filename:
                filename = DataExporter._generate_filename(prefix, "xlsx")
            
            # Garantir que termina com .xlsx
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Caminho completo
            export_dir = DataExporter.get_export_dir()
            filepath = os.path.join(export_dir, filename)
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Estilo do cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Escrever cabeçalho
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Escrever dados
            for row_num, row_data in enumerate(data, 2):
                for col_num, value in enumerate(row_data, 1):
                    # Converter None para string vazia
                    cell_value = value if value is not None else ''
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salvar
            wb.save(filepath)
            
            return True, filepath
            
        except Exception as e:
            return False, f"Erro ao exportar Excel: {str(e)}"
    
    @staticmethod
    def export_pedidos_csv(pedidos: List[Tuple]) -> Tuple[bool, str]:
        """
        Exporta pedidos para CSV.
        
        Args:
            pedidos: Lista de tuplas com dados dos pedidos
            
        Returns:
            Tupla (sucesso: bool, caminho_ou_mensagem: str)
        """
        headers = [
            "ID", "Número OS", "Data Criação", "Cliente", "CPF", "Telefone",
            "Detalhes Produto", "Valor Produto", "Valor Entrada", "Frete",
            "Forma Pagamento", "Prazo", "Nome PDF", "Status"
        ]
        return DataExporter.export_to_csv(pedidos, headers, prefix="pedidos")
    
    @staticmethod
    def export_pedidos_excel(pedidos: List[Tuple]) -> Tuple[bool, str]:
        """
        Exporta pedidos para Excel.
        
        Args:
            pedidos: Lista de tuplas com dados dos pedidos
            
        Returns:
            Tupla (sucesso: bool, caminho_ou_mensagem: str)
        """
        headers = [
            "ID", "Número OS", "Data Criação", "Cliente", "CPF", "Telefone",
            "Detalhes Produto", "Valor Produto", "Valor Entrada", "Frete",
            "Forma Pagamento", "Prazo", "Nome PDF", "Status"
        ]
        return DataExporter.export_to_excel(pedidos, headers, prefix="pedidos", 
                                           sheet_name="Ordens de Serviço")
    
    @staticmethod
    def export_clientes_csv(clientes: List[Tuple]) -> Tuple[bool, str]:
        """Exporta clientes para CSV"""
        headers = [
            "ID", "Nome", "CPF", "CNPJ", "Inscrição Estadual", "Telefone", "Email",
            "CEP", "Rua", "Número", "Bairro", "Cidade", "UF", "Referência", "Data Criação"
        ]
        return DataExporter.export_to_csv(clientes, headers, prefix="clientes")
    
    @staticmethod
    def export_clientes_excel(clientes: List[Tuple]) -> Tuple[bool, str]:
        """Exporta clientes para Excel"""
        headers = [
            "ID", "Nome", "CPF", "CNPJ", "Inscrição Estadual", "Telefone", "Email",
            "CEP", "Rua", "Número", "Bairro", "Cidade", "UF", "Referência", "Data Criação"
        ]
        return DataExporter.export_to_excel(clientes, headers, prefix="clientes", 
                                           sheet_name="Clientes")
    
    @staticmethod
    def export_produtos_csv(produtos: List[Tuple]) -> Tuple[bool, str]:
        """Exporta produtos para CSV"""
        headers = [
            "ID", "Nome", "Código", "Preço", "Descrição", "Categoria", "Data Criação"
        ]
        return DataExporter.export_to_csv(produtos, headers, prefix="produtos")
    
    @staticmethod
    def export_produtos_excel(produtos: List[Tuple]) -> Tuple[bool, str]:
        """Exporta produtos para Excel"""
        headers = [
            "ID", "Nome", "Código", "Preço", "Descrição", "Categoria", "Data Criação"
        ]
        return DataExporter.export_to_excel(produtos, headers, prefix="produtos", 
                                           sheet_name="Produtos")
    
    @staticmethod
    def abrir_pasta_exportacoes():
        """Abre a pasta de exportações no explorador de arquivos"""
        import platform
        import subprocess
        
        export_dir = DataExporter.get_export_dir()
        
        try:
            if platform.system() == "Windows":
                os.startfile(export_dir)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", export_dir])
            else:  # Linux
                subprocess.run(["xdg-open", export_dir])
            return True
        except Exception as e:
            print(f"Erro ao abrir pasta: {e}")
            return False
